"""Helpers for converting model data to/from XLSX files."""

from __future__ import annotations

from io import BytesIO
from typing import Iterable, List, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


def build_xlsx(headers: Sequence[str], rows: Iterable[Sequence[object]]) -> bytes:
    """Return XLSX binary with provided headers and rows."""

    wb = Workbook()
    ws = wb.active
    ws.append(list(headers))
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        values = [_normalize_cell(value) for value in row]
        ws.append(values)
        row_idx = ws.max_row
        if row_idx % 2 == 0:
            stripe = PatternFill(start_color="F4F6FD", end_color="F4F6FD", fill_type="solid")
            for cell in ws[row_idx]:
                cell.fill = stripe

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_xlsx(data: bytes) -> List[dict[str, object]]:
    """Parse XLSX binary into list of rows keyed by headers."""

    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    headers: List[str] = []
    rows: List[dict[str, object]] = []

    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [cell for cell in row]
        if idx == 1:
            headers = [_safe_header(cell) for cell in values if _safe_header(cell)]
            continue
        if not headers:
            break
        row_data: dict[str, object] = {}
        for header, cell in zip(headers, values):
            row_data[header] = cell
        if any(value not in (None, "") for value in row_data.values()):
            rows.append(row_data)
    return rows


def _normalize_cell(value: object) -> object:
    if value is None:
        return ""
    return value


def _safe_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
